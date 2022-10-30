from selenium.webdriver import Firefox
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import tkinter as tk
from tkinter import filedialog
from selenium import webdriver


url = "https://www.ibs.it/"

root = tk.Tk()
root.withdraw()
file = filedialog.askopenfilename()

driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
driver.maximize_window()
driver.get(url)

wb = openpyxl.load_workbook(file)
sh = wb.active
m_row = sh.max_row

def split_title(cell):
    string1 = sh.cell(row = i, column=2).value.split(" / ")
    string2 = string1[0].replace("*", "")
    return string2

def get_abstract(i):
    driver.get(url)
    txt = split_title(i)
    driver.find_element(By.ID, 'inputSearch').send_keys(
        txt)
    sleep(1)
    driver.find_element(By.CLASS_NAME, 'cc-search-button').click()
    sleep(1)
    driver.find_element(By.XPATH, 
    '//*[@id="search-body"]/div[1]/div/div/div[3]/div[1]/div/div[2]/div[3]/div[1]/div[3]/div/a').click()
    sleep(1)
    t = driver.find_element(By.XPATH, 
    '//*[@id="pdpMainRow"]/div[2]/div/div[1]/h1').get_attribute("innerHTML").strip().replace("\' ", "\'").split()

    txtl = txt.split()
    ts = t.sort()
    txts = txtl.sort()

    if(ts == txts):
        ab = driver.find_element(By.XPATH,
        '//*[@id="pdp-descrizione"]/div/div/div[2]/div[1]')
        abText = ab.get_attribute("innerHTML").replace("<br>", "").strip()
        abCell = sh.cell(row = i, column=4)
        abCell.value = abText
        wb.save(file)
    else:
        pass

for i in range(1, m_row + 1):
    try:
        get_abstract(i)
    except:
        pass