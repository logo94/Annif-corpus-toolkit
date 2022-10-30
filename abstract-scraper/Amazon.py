from selenium.webdriver import Firefox
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import tkinter as tk
from tkinter import filedialog


url = "https://www.amazon.it/libri-italiano/b?ie=UTF8&node=411663031"

root = tk.Tk()
root.withdraw()
file = filedialog.askopenfilename()

fire_driver = GeckoDriverManager().install()
driver = Firefox(service=FirefoxService(fire_driver))
driver.maximize_window()
driver.get(url)

wb = openpyxl.load_workbook(file)
sh = wb.active
m_row = sh.max_row

def split_title(cell):
    string1 = sh.cell(row = i, column=2).value.split(" / ")
    string2 = string1[0].replace("*", "")
    return string2

def get_abstract(i):
    driver.get(url)
    txt = split_title(i)
    driver.find_element(By.ID, 'twotabsearchtextbox').send_keys(
        txt)
    sleep(0.5)
    driver.find_element(By.ID, 'nav-search-submit-button').click()
    sleep(1)
    driver.find_element(By.XPATH, 
    '/html/body/div[1]/div[2]/div[1]/div[1]/div/span[3]/div[2]/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/h2/a/span').click()
    sleep(0.5)
    t = driver.find_element(By.XPATH,
    '//*[@id="productTitle"]').get_attribute("innerHTML").strip().replace("\' ", "\'").split()

    txtl = txt.split()
    ts = t.sort()
    txts = txtl.sort()

    if(ts == txts):
        ab = driver.find_element(By.XPATH,
        '/html/body/div[2]/div[2]/div[4]/div[1]/div[7]/div[29]/div/div[1]/span')
        abText = ab.get_attribute("innerHTML").replace("<br>", "").strip()
        abCell = sh.cell(row = i, column=4)
        abCell.value = abText
        wb.save(file)
    else:
        pass


for i in range(1, m_row + 1):
    try:
        get_abstract(i)
    except:
        pass
