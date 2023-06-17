from operator import contains
from time import sleep
import openpyxl
import tkinter as tk
from tkinter import filedialog
import json
import requests

# Selezione del foglio di calcolo
root = tk.Tk()
root.withdraw()
file = filedialog.askopenfilename()

# Apertura del foglio di calcolo per la lettura e la scrittura
wb = openpyxl.load_workbook(file)
sh = wb.active
m_row = sh.max_row

# Interrogazione API servizio di riconciliazione Wikidata
def get_uri(title):
    query = {
        "q0": {
            "query": title,
            "limit": 5,
            "type_strict": "should"
        }
    }
    http = requests.Session()
    payload = {'queries': json.dumps(query)}
    response = http.post('https://wikidata.reconci.link/it/api', data=payload).json()
    try:
        for candidate in response['q0']['result']:
            if candidate['match'] == True:
                wikidata_id = "<https://www.wikidata.org/wiki/" + candidate['id'] + ">"
    except:
        raise Exception('URI not found')
    return wikidata_id

# Per ogni riga del foglio di calcolo viene letta l'etichetta e viene assegnato un URI
for i in range(1, m_row + 1):
    try:
        title = sh.cell(row = i, column=2).value
        wiki_uri = get_uri(i, title)
        sh.cell(row = i, column=1).value = wiki_uri
        wb.save(file)
    except:
        continue
